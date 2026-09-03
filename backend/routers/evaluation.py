from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from pydantic import BaseModel
from datetime import datetime
from typing import Optional

from database.connection import get_db, SessionLocal
from models.company import Company, CompanyMember
from models.team import Team, TeamMember
from models.evaluation import EvaluatorAssignment, EvaluationCycle, EvaluationEntry, EvaluationResult, OneOnOneSession
from routers.deps import get_current_user
from utils.push import send_push_to_users
from utils.meeting_ai import transcribe_audio, analyze_career, analyze_growth, analyze_one_on_one
from utils.admin import is_superadmin_email

router = APIRouter()

CATEGORIES = ("performance", "competency")


# ── 공통 권한 헬퍼 ─────────────────────────────────────────
def _get_member(db: Session, uid: str, company_id: str) -> Optional[CompanyMember]:
    return db.query(CompanyMember).filter(
        CompanyMember.user_id == uid, CompanyMember.company_id == company_id,
    ).first()


def _get_primary_member(db: Session, uid: str) -> Optional[CompanyMember]:
    """uid만으로 소속 회사를 찾을 때 쓰는 조회. 한 계정이 여러 회사에 중복 등록된 경우
    ORDER BY 없는 .first()는 호출마다 다른 행을 돌려줄 수 있어(같은 화면에서 값이
    널뛰는 버그의 원인이었음) 항상 "가장 최근에 등록된 소속"으로 고정한다."""
    return (
        db.query(CompanyMember)
        .filter(CompanyMember.user_id == uid)
        .order_by(CompanyMember.created_at.desc())
        .first()
    )


def _require_member(db: Session, uid: str, company_id: str) -> CompanyMember:
    member = _get_member(db, uid, company_id)
    if not member:
        raise HTTPException(status_code=403, detail="해당 회사 소속만 이용할 수 있어요")
    return member


def _require_admin(db: Session, current_user: dict, company_id: str):
    if is_superadmin_email(db, current_user.get("email")):
        return
    member = _get_member(db, current_user["uid"], company_id)
    if not member or not member.is_admin:
        raise HTTPException(status_code=403, detail="관리자만 이용할 수 있어요")


def _require_reviewer(db: Session, current_user: dict, entry: EvaluationEntry):
    if is_superadmin_email(db, current_user.get("email")) or entry.evaluator_id == current_user["uid"]:
        return
    member = _get_member(db, current_user["uid"], entry.company_id)
    if not member or not member.is_admin:
        raise HTTPException(status_code=403, detail="평가자 또는 관리자만 검토할 수 있어요")


def _name_map(db: Session, company_id: str) -> dict:
    members = db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    return {m.user_id: (m.user_name or m.user_email) for m in members}


# ── 직렬화 ─────────────────────────────────────────────────
def _serialize_cycle(c: EvaluationCycle) -> dict:
    return {
        "id": c.id,
        "code": c.code,
        "name": c.name,
        "plan_start": c.plan_start,
        "plan_end": c.plan_end,
        "actual_start": c.actual_start,
        "actual_end": c.actual_end,
        "review_start": c.review_start,
        "review_end": c.review_end,
        "grade_distribution": c.grade_distribution or [],
        "status": c.status,
    }


def _serialize_entry(e: EvaluationEntry, names: dict | None = None) -> dict:
    names = names or {}
    return {
        "id": e.id,
        "cycle_id": e.cycle_id,
        "user_id": e.user_id,
        "user_name": names.get(e.user_id, e.user_id),
        "evaluator_id": e.evaluator_id,
        "evaluator_name": names.get(e.evaluator_id, e.evaluator_id),
        "category": e.category,
        "plan_content": e.plan_content,
        "plan_status": e.plan_status,
        "plan_feedback": e.plan_feedback,
        "plan_submitted_at": e.plan_submitted_at.isoformat() if e.plan_submitted_at else None,
        "actual_content": e.actual_content,
        "actual_status": e.actual_status,
        "actual_feedback": e.actual_feedback,
        "actual_submitted_at": e.actual_submitted_at.isoformat() if e.actual_submitted_at else None,
    }


# ── 화면 진입 시 필요한 데이터를 한 번에 묶어 반환 (요청 왕복 최소화) ──
def _get_active_cycle_row(db: Session, company_id: str) -> Optional[EvaluationCycle]:
    return (
        db.query(EvaluationCycle)
        .filter(EvaluationCycle.company_id == company_id, EvaluationCycle.status == "active")
        .order_by(EvaluationCycle.created_at.desc())
        .first()
    )


@router.get("/bootstrap")
def get_bootstrap(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """/evaluation 메인 화면용. company/my + cycles/active + entries/my + results/me를
    하나로 묶어 화면 진입 시 요청 왕복 수를 4회 → 1회로 줄인다."""
    uid = current_user["uid"]
    member = _get_primary_member(db, uid)
    company_id = member.company_id if member else None

    cycle_data = None
    entries_data: list[dict] = []
    result_data = _serialize_result(None)

    if company_id:
        cycle_row = _get_active_cycle_row(db, company_id)
        if cycle_row:
            cycle_data = _serialize_cycle(cycle_row)
            entry_rows = db.query(EvaluationEntry).filter(
                EvaluationEntry.cycle_id == cycle_row.id, EvaluationEntry.user_id == uid,
            ).all()
            names = _name_map(db, company_id) if entry_rows else {}
            entries_data = [_serialize_entry(e, names) for e in entry_rows]
            result_row = db.query(EvaluationResult).filter(
                EvaluationResult.cycle_id == cycle_row.id, EvaluationResult.user_id == uid,
            ).first()
            result_data = _serialize_result(result_row)

    return {
        "company_id": company_id,
        "is_manager": bool(member and member.is_manager),
        "job_title": member.job_title if member else None,
        "cycle": cycle_data,
        "entries": entries_data,
        "result": result_data,
    }


@router.get("/bootstrap/review")
def get_bootstrap_review(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """/evaluation/review 화면용. company/my + cycles/active + entries/review + results를
    하나로 묶어 요청 왕복 수를 4회 → 1회로 줄인다."""
    uid = current_user["uid"]
    member = _get_primary_member(db, uid)
    company_id = member.company_id if member else None

    cycle_data = None
    entries_data: list[dict] = []
    results_payload = {"people": [], "distribution": []}

    if company_id:
        cycle_row = _get_active_cycle_row(db, company_id)
        if cycle_row:
            cycle_data = _serialize_cycle(cycle_row)
            entry_rows = db.query(EvaluationEntry).filter(
                EvaluationEntry.cycle_id == cycle_row.id, EvaluationEntry.evaluator_id == uid,
            ).all()
            names = _name_map(db, company_id) if entry_rows else {}
            entries_data = [_serialize_entry(e, names) for e in entry_rows]
            results_payload = _compute_results(db, cycle_row, current_user)

    return {"company_id": company_id, "cycle": cycle_data, "entries": entries_data, **results_payload}


@router.get("/bootstrap/settings")
def get_bootstrap_settings(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """/evaluation/settings 화면용. company/my + members + cycles + teams를 하나로 묶어
    요청 왕복 수를 줄인다. 평가자 매핑은 이제 사이클(평가코드)별로 관리되므로 여기 포함하지
    않고, 특정 사이클을 선택했을 때 GET /assignments/{cycle_id}로 따로 조회한다."""
    uid = current_user["uid"]
    member = _get_primary_member(db, uid)
    company_id = member.company_id if member else None
    if not company_id:
        return {"company_id": None, "evaluation_enabled": False, "members": [], "cycles": [], "teams": []}

    company = db.query(Company).filter(Company.id == company_id).first()
    members = db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    cycles = (
        db.query(EvaluationCycle)
        .filter(EvaluationCycle.company_id == company_id)
        .order_by(EvaluationCycle.created_at.desc())
        .all()
    )
    teams = db.query(Team).filter(Team.company_id == company_id).all()

    names = {m.user_id: (m.user_name or m.user_email) for m in members}

    return {
        "company_id": company_id,
        "evaluation_enabled": bool(company and company.evaluation_enabled),
        "members": [
            {"user_id": m.user_id, "user_name": m.user_name, "user_email": m.user_email, "org_level": m.org_level}
            for m in members
        ],
        "cycles": [_serialize_cycle(c) for c in cycles],
        "teams": [
            {
                "id": t.id, "name": t.name, "manager_id": t.manager_id,
                "manager_name": names.get(t.manager_id) if t.manager_id else None,
                "parent_team_id": t.parent_team_id,
            }
            for t in teams
        ],
    }


# ── 평가 기능 on/off ───────────────────────────────────────
class ToggleEvaluationRequest(BaseModel):
    evaluation_enabled: bool


@router.put("/toggle/{company_id}")
def toggle_evaluation(
    company_id: str, req: ToggleEvaluationRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없어요")
    _require_admin(db, current_user, company_id)
    company.evaluation_enabled = req.evaluation_enabled
    company.updated_by = current_user["uid"]
    db.commit()
    return {"success": True, "evaluation_enabled": req.evaluation_enabled}


# ── 평가자 매핑 설정 (사이클별) ─────────────────────────────
def _get_cycle_or_404(db: Session, cycle_id: str) -> EvaluationCycle:
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    return cycle


@router.post("/assignments/seed/{cycle_id}")
def seed_assignments(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    """조직도(팀장) 기준으로 이 사이클에 매핑이 없는 사람만 자동 생성. 이미 있는 매핑은 건드리지 않는다."""
    cycle = _get_cycle_or_404(db, cycle_id)
    _require_admin(db, current_user, cycle.company_id)

    teams = db.query(Team).filter(Team.company_id == cycle.company_id).all()
    manager_by_team = {t.id: t.manager_id for t in teams if t.manager_id}
    if not manager_by_team:
        return {"success": True, "created": 0}

    memberships = db.query(TeamMember).filter(TeamMember.team_id.in_(manager_by_team.keys())).all()
    existing = {
        r.evaluatee_user_id
        for r in db.query(EvaluatorAssignment).filter(EvaluatorAssignment.cycle_id == cycle_id).all()
    }

    created = 0
    for tm in memberships:
        if tm.user_id in existing:
            continue
        manager_id = manager_by_team.get(tm.team_id)
        if not manager_id or manager_id == tm.user_id:
            continue
        db.add(EvaluatorAssignment(
            company_id=cycle.company_id, cycle_id=cycle_id, evaluatee_user_id=tm.user_id, evaluator_user_id=manager_id,
            source="auto", created_by=current_user["uid"],
        ))
        existing.add(tm.user_id)
        created += 1
    db.commit()
    return {"success": True, "created": created}


@router.get("/assignments/{cycle_id}")
def list_assignments(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = _get_cycle_or_404(db, cycle_id)
    _require_admin(db, current_user, cycle.company_id)
    rows = db.query(EvaluatorAssignment).filter(EvaluatorAssignment.cycle_id == cycle_id).all()
    names = _name_map(db, cycle.company_id)
    return {
        "assignments": [
            {
                "id": r.id,
                "evaluatee_user_id": r.evaluatee_user_id,
                "evaluatee_name": names.get(r.evaluatee_user_id, r.evaluatee_user_id),
                "evaluator_user_id": r.evaluator_user_id,
                "evaluator_name": names.get(r.evaluator_user_id, r.evaluator_user_id),
                "source": r.source,
            }
            for r in rows
        ]
    }


class AssignmentUpsertRequest(BaseModel):
    cycle_id: str
    evaluatee_user_id: str
    evaluator_user_id: str


@router.put("/assignments")
def upsert_assignment(
    req: AssignmentUpsertRequest, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = _get_cycle_or_404(db, req.cycle_id)
    _require_admin(db, current_user, cycle.company_id)
    if req.evaluatee_user_id == req.evaluator_user_id:
        raise HTTPException(status_code=400, detail="본인을 평가자로 지정할 수 없어요")

    row = db.query(EvaluatorAssignment).filter(
        EvaluatorAssignment.cycle_id == req.cycle_id,
        EvaluatorAssignment.evaluatee_user_id == req.evaluatee_user_id,
    ).first()
    if row:
        row.evaluator_user_id = req.evaluator_user_id
        row.source = "manual"
        row.updated_by = current_user["uid"]
    else:
        row = EvaluatorAssignment(
            company_id=cycle.company_id, cycle_id=req.cycle_id, evaluatee_user_id=req.evaluatee_user_id,
            evaluator_user_id=req.evaluator_user_id, source="manual", created_by=current_user["uid"],
        )
        db.add(row)
    db.commit()
    db.refresh(row)
    return {"success": True, "id": row.id}


@router.delete("/assignments/{assignment_id}")
def delete_assignment(
    assignment_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    row = db.query(EvaluatorAssignment).filter(EvaluatorAssignment.id == assignment_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="매핑을 찾을 수 없어요")
    _require_admin(db, current_user, row.company_id)
    db.delete(row)
    db.commit()
    return {"success": True}


# ── 평가자 설정 엑셀 양식 다운로드 / 업로드 ──────────────────
ASSIGNMENT_TEMPLATE_HEADERS = ["이름", "이메일", "레벨(숫자)", "상위자(평가자) 이메일"]
ASSIGNMENT_TEMPLATE_MAX_ROWS = 2000
ASSIGNMENT_TEMPLATE_MAX_BYTES = 3 * 1024 * 1024  # 3MB


@router.get("/assignments/template/{cycle_id}")
def download_assignment_template(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    """조직도 계층 구조(이름/이메일/레벨/상위자 이메일)를 입력할 엑셀 양식을 내려준다.
    기존 직원들의 이름/이메일은 미리 채워주고, 레벨/상위자 이메일만 입력하면 되게 한다."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    import io

    cycle = _get_cycle_or_404(db, cycle_id)
    _require_admin(db, current_user, cycle.company_id)
    members = db.query(CompanyMember).filter(CompanyMember.company_id == cycle.company_id).order_by(CompanyMember.created_at).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "평가자 설정"

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(ASSIGNMENT_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 24

    for row_idx, m in enumerate(members, 2):
        ws.cell(row=row_idx, column=1, value=m.user_name or "")
        ws.cell(row=row_idx, column=2, value=m.user_email)
        ws.cell(row=row_idx, column=3, value=m.org_level)

    # 레벨 칸에 정수만 입력하도록 유효성 검사(엑셀 자체 UX 가드 — 서버 검증을 대체하진 않음)
    if members:
        dv = DataValidation(type="whole", operator="greaterThan", formula1="0", showErrorMessage=True)
        dv.error = "레벨은 1 이상의 숫자로 입력해주세요"
        ws.add_data_validation(dv)
        dv.add(f"C2:C{len(members) + 1}")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "평가자_설정_양식.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/assignments/upload/{cycle_id}")
async def upload_assignments(
    cycle_id: str, file: UploadFile = File(...),
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    """엑셀 양식을 업로드해 조직도(레벨/상위자)를 일괄 반영한다. 이 사이클의 기존 평가자
    매핑은 전체 교체된다(다른 사이클의 매핑에는 영향 없음).
    업로드 취약점 방어: 확장자/크기/실제 파일 시그니처(zip)/헤더 형식을 모두 검사하고,
    우리 양식과 다르면 이유를 명시해 거부한다."""
    import openpyxl
    import io

    cycle = _get_cycle_or_404(db, cycle_id)
    company_id = cycle.company_id
    _require_admin(db, current_user, company_id)

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="지정된 엑셀 양식(.xlsx)만 업로드할 수 있어요")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="파일이 비어있어요")
    if len(content) > ASSIGNMENT_TEMPLATE_MAX_BYTES:
        raise HTTPException(status_code=400, detail="파일이 너무 커요 (최대 3MB)")
    # xlsx는 ZIP 컨테이너 포맷 — 실제 파일 시그니처를 확인해 확장자만 바꾼 다른 파일을 차단한다.
    if content[:2] != b"PK":
        raise HTTPException(status_code=400, detail="올바른 엑셀(.xlsx) 파일이 아니에요")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="파일을 열 수 없어요. 지정된 양식 파일이 맞는지 확인해주세요")

    ws = wb.active
    header_row = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if header_row[: len(ASSIGNMENT_TEMPLATE_HEADERS)] != ASSIGNMENT_TEMPLATE_HEADERS:
        raise HTTPException(
            status_code=400,
            detail="지정된 양식이 아니에요. '템플릿 다운로드'로 받은 파일을 그대로 사용해주세요",
        )

    members = db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    member_by_email = {m.user_email.strip().lower(): m for m in members if m.user_email}

    rows = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if i - 1 > ASSIGNMENT_TEMPLATE_MAX_ROWS:
            raise HTTPException(status_code=400, detail=f"한 번에 최대 {ASSIGNMENT_TEMPLATE_MAX_ROWS}행까지 처리할 수 있어요")
        name, email, level, manager_email = (raw + (None, None, None, None))[:4]
        if not email or not str(email).strip():
            continue
        rows.append({
            "row": i, "name": str(name).strip() if name else "",
            "email": str(email).strip().lower(),
            "level": level, "manager_email": str(manager_email).strip().lower() if manager_email else None,
        })

    if not rows:
        raise HTTPException(status_code=400, detail="입력된 데이터가 없어요")

    errors = []
    for r in rows:
        if r["email"] not in member_by_email:
            errors.append(f"{r['row']}행: '{r['email']}'은 이 회사에 등록된 이메일이 아니에요")
            continue
        if r["level"] is not None:
            try:
                lvl = int(r["level"])
                if lvl <= 0:
                    raise ValueError()
            except (TypeError, ValueError):
                errors.append(f"{r['row']}행: 레벨은 1 이상의 숫자여야 해요 (입력값: {r['level']!r})")
        if r["manager_email"]:
            if r["manager_email"] == r["email"]:
                errors.append(f"{r['row']}행: 본인을 상위자로 지정할 수 없어요")
            elif r["manager_email"] not in member_by_email:
                errors.append(f"{r['row']}행: 상위자 이메일 '{r['manager_email']}'을 찾을 수 없어요")

    if errors:
        raise HTTPException(status_code=400, detail={"message": "업로드 내용에 오류가 있어요", "errors": errors[:50]})

    # 검증 통과 → org_level 반영 + 평가자 매핑 전체 교체
    for r in rows:
        member = member_by_email[r["email"]]
        member.org_level = int(r["level"]) if r["level"] is not None else None
        member.updated_by = current_user["uid"]

    db.query(EvaluatorAssignment).filter(EvaluatorAssignment.cycle_id == cycle_id).delete()
    created = 0
    for r in rows:
        if not r["manager_email"]:
            continue
        evaluatee = member_by_email[r["email"]]
        evaluator = member_by_email[r["manager_email"]]
        db.add(EvaluatorAssignment(
            company_id=company_id, cycle_id=cycle_id, evaluatee_user_id=evaluatee.user_id,
            evaluator_user_id=evaluator.user_id, source="excel", created_by=current_user["uid"],
        ))
        created += 1

    db.commit()
    return {"success": True, "processed": len(rows), "assignments_created": created}


# ── 평가 코드(기준정보) ────────────────────────────────────
class CycleCreateRequest(BaseModel):
    company_id: str
    code: str
    name: str


class CycleUpdateRequest(BaseModel):
    name: Optional[str] = None
    plan_start: Optional[str] = None
    plan_end: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    review_start: Optional[str] = None
    review_end: Optional[str] = None
    grade_distribution: Optional[list[dict]] = None


@router.post("/cycles")
def create_cycle(
    req: CycleCreateRequest, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    _require_admin(db, current_user, req.company_id)
    if not req.code.strip() or not req.name.strip():
        raise HTTPException(status_code=400, detail="평가 코드와 이름을 입력해주세요")
    cycle = EvaluationCycle(
        company_id=req.company_id, code=req.code.strip(), name=req.name.strip(),
        created_by=current_user["uid"],
    )
    db.add(cycle)
    db.commit()
    db.refresh(cycle)
    return _serialize_cycle(cycle)


@router.put("/cycles/{cycle_id}")
def update_cycle(
    cycle_id: str, req: CycleUpdateRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    _require_admin(db, current_user, cycle.company_id)
    if cycle.status != "draft":
        raise HTTPException(status_code=400, detail="이미 시작된 평가는 기준정보를 수정할 수 없어요")

    if req.name is not None:
        cycle.name = req.name.strip()
    for field in ("plan_start", "plan_end", "actual_start", "actual_end", "review_start", "review_end"):
        val = getattr(req, field)
        if val is not None:
            setattr(cycle, field, val)
    if req.grade_distribution is not None:
        cycle.grade_distribution = req.grade_distribution
    cycle.updated_by = current_user["uid"]
    db.commit()
    db.refresh(cycle)
    return _serialize_cycle(cycle)


@router.delete("/cycles/{cycle_id}")
def delete_cycle(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    _require_admin(db, current_user, cycle.company_id)
    if cycle.status != "draft":
        raise HTTPException(status_code=400, detail="이미 시작된 평가는 삭제할 수 없어요")

    db.query(EvaluatorAssignment).filter(EvaluatorAssignment.cycle_id == cycle_id).delete()
    db.delete(cycle)
    db.commit()
    return {"success": True}


@router.get("/cycles/active/{company_id}")
def get_active_cycle(
    company_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    _require_member(db, current_user["uid"], company_id)
    cycle = (
        db.query(EvaluationCycle)
        .filter(EvaluationCycle.company_id == company_id, EvaluationCycle.status == "active")
        .order_by(EvaluationCycle.created_at.desc())
        .first()
    )
    return {"cycle": _serialize_cycle(cycle) if cycle else None}


@router.get("/cycles/{company_id}")
def list_cycles(
    company_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    _require_admin(db, current_user, company_id)
    cycles = (
        db.query(EvaluationCycle)
        .filter(EvaluationCycle.company_id == company_id)
        .order_by(EvaluationCycle.created_at.desc())
        .all()
    )
    return {"cycles": [_serialize_cycle(c) for c in cycles]}


@router.post("/cycles/{cycle_id}/activate")
def activate_cycle(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    _require_admin(db, current_user, cycle.company_id)
    if cycle.status != "draft":
        raise HTTPException(status_code=400, detail="이미 시작됐거나 종료된 평가예요")

    periods = [cycle.plan_start, cycle.plan_end, cycle.actual_start, cycle.actual_end, cycle.review_start, cycle.review_end]
    if any(not p for p in periods):
        raise HTTPException(status_code=400, detail="계획/실적/평가 입력기간을 모두 설정해야 시작할 수 있어요")
    if not (cycle.plan_start <= cycle.plan_end <= cycle.actual_start <= cycle.actual_end <= cycle.review_start <= cycle.review_end):
        raise HTTPException(status_code=400, detail="기간 순서가 올바르지 않아요 (계획 → 실적 → 평가 순으로 설정해주세요)")

    dist = cycle.grade_distribution or []
    if not dist:
        raise HTTPException(status_code=400, detail="등급별 비율을 설정해야 시작할 수 있어요")
    total_ratio = sum(g.get("ratio", 0) for g in dist)
    if abs(total_ratio - 100) > 0.01:
        raise HTTPException(status_code=400, detail=f"등급별 비율의 합이 100이어야 해요 (현재 {total_ratio})")

    assignments = db.query(EvaluatorAssignment).filter(EvaluatorAssignment.cycle_id == cycle.id).all()
    if not assignments:
        raise HTTPException(status_code=400, detail="이 사이클에 평가자 매핑이 없어요. 먼저 평가자 설정을 완료해주세요")

    created = 0
    for a in assignments:
        for category in CATEGORIES:
            exists = db.query(EvaluationEntry).filter(
                EvaluationEntry.cycle_id == cycle.id,
                EvaluationEntry.user_id == a.evaluatee_user_id,
                EvaluationEntry.category == category,
            ).first()
            if exists:
                continue
            db.add(EvaluationEntry(
                cycle_id=cycle.id, company_id=cycle.company_id,
                user_id=a.evaluatee_user_id, evaluator_id=a.evaluator_user_id, category=category,
            ))
            created += 1

    cycle.status = "active"
    cycle.updated_by = current_user["uid"]
    db.commit()
    return {"success": True, "status": "active", "entries_created": created}


# ── 계획/실적 작성 (피평가자) ──────────────────────────────
class EntryContentRequest(BaseModel):
    content: str
    submit: bool = False


@router.get("/entries/my/{cycle_id}")
def my_entries(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entries = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == cycle_id, EvaluationEntry.user_id == current_user["uid"],
    ).all()
    names = _name_map(db, entries[0].company_id) if entries else {}
    return {"entries": [_serialize_entry(e, names) for e in entries]}


@router.put("/entries/{entry_id}/plan")
def write_plan(
    entry_id: str, req: EntryContentRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entry = db.query(EvaluationEntry).filter(EvaluationEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    if entry.user_id != current_user["uid"]:
        raise HTTPException(status_code=403, detail="본인의 평가만 작성할 수 있어요")
    if entry.plan_status == "approved":
        raise HTTPException(status_code=400, detail="이미 승인된 계획이에요")

    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == entry.cycle_id).first()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    if not cycle or not (cycle.plan_start and cycle.plan_end and cycle.plan_start <= today <= cycle.plan_end):
        raise HTTPException(status_code=400, detail="계획 입력기간이 아니에요")

    entry.plan_content = req.content
    entry.plan_status = "submitted" if req.submit else "draft"
    if req.submit:
        entry.plan_submitted_at = datetime.utcnow()
        entry.plan_feedback = None  # 재제출 시 이전 회차 피드백은 정리
    db.commit()

    if req.submit:
        try:
            names = _name_map(db, entry.company_id)
            send_push_to_users(
                db, [entry.evaluator_id], title="📝 평가 계획 제출",
                body=f"{names.get(entry.user_id, entry.user_id)}님이 계획을 제출했어요.",
                url="/evaluation/review",
            )
        except Exception as e:
            print(f"[write_plan] 알림 전송 실패: {e}")

    return _serialize_entry(entry)


@router.put("/entries/{entry_id}/actual")
def write_actual(
    entry_id: str, req: EntryContentRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entry = db.query(EvaluationEntry).filter(EvaluationEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    if entry.user_id != current_user["uid"]:
        raise HTTPException(status_code=403, detail="본인의 평가만 작성할 수 있어요")
    if entry.plan_status != "approved":
        raise HTTPException(status_code=400, detail="계획이 먼저 승인돼야 실적을 입력할 수 있어요")
    if entry.actual_status == "approved":
        raise HTTPException(status_code=400, detail="이미 승인된 실적이에요")

    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == entry.cycle_id).first()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    if not cycle or not (cycle.actual_start and cycle.actual_end and cycle.actual_start <= today <= cycle.actual_end):
        raise HTTPException(status_code=400, detail="실적 입력기간이 아니에요")

    entry.actual_content = req.content
    entry.actual_status = "submitted" if req.submit else "draft"
    if req.submit:
        entry.actual_submitted_at = datetime.utcnow()
        entry.actual_feedback = None  # 재제출 시 이전 회차 피드백은 정리
    db.commit()

    if req.submit:
        try:
            names = _name_map(db, entry.company_id)
            send_push_to_users(
                db, [entry.evaluator_id], title="📝 평가 실적 제출",
                body=f"{names.get(entry.user_id, entry.user_id)}님이 실적을 제출했어요.",
                url="/evaluation/review",
            )
        except Exception as e:
            print(f"[write_actual] 알림 전송 실패: {e}")

    return _serialize_entry(entry)


# ── 검토/승인 (평가자) ─────────────────────────────────────
class ReviewRequest(BaseModel):
    status: str  # approved / feedback
    feedback: str = ""


@router.get("/entries/review/{cycle_id}")
def review_entries(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entries = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == cycle_id, EvaluationEntry.evaluator_id == current_user["uid"],
    ).all()
    names = _name_map(db, entries[0].company_id) if entries else {}
    return {"entries": [_serialize_entry(e, names) for e in entries]}


@router.put("/entries/{entry_id}/plan/review")
def review_plan(
    entry_id: str, req: ReviewRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entry = db.query(EvaluationEntry).filter(EvaluationEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    _require_reviewer(db, current_user, entry)
    if req.status not in ("approved", "feedback"):
        raise HTTPException(status_code=400, detail="status는 approved 또는 feedback만 가능해요")
    if entry.plan_status != "submitted":
        raise HTTPException(status_code=400, detail="제출된 계획만 검토할 수 있어요")

    entry.plan_status = req.status
    entry.plan_feedback = req.feedback
    entry.plan_reviewed_at = datetime.utcnow()
    entry.plan_reviewed_by = current_user["uid"]
    db.commit()

    status_text = "승인" if req.status == "approved" else "피드백 반영 요청"
    try:
        send_push_to_users(
            db, [entry.user_id], title=f"📋 평가 계획 {status_text}",
            body=req.feedback or f"계획이 {status_text}됐어요.", url="/evaluation",
        )
    except Exception as e:
        print(f"[review_plan] 알림 전송 실패: {e}")

    return _serialize_entry(entry)


@router.put("/entries/{entry_id}/actual/review")
def review_actual(
    entry_id: str, req: ReviewRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    entry = db.query(EvaluationEntry).filter(EvaluationEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    _require_reviewer(db, current_user, entry)
    if req.status not in ("approved", "feedback"):
        raise HTTPException(status_code=400, detail="status는 approved 또는 feedback만 가능해요")
    if entry.actual_status != "submitted":
        raise HTTPException(status_code=400, detail="제출된 실적만 검토할 수 있어요")

    entry.actual_status = req.status
    entry.actual_feedback = req.feedback
    entry.actual_reviewed_at = datetime.utcnow()
    entry.actual_reviewed_by = current_user["uid"]
    db.commit()

    status_text = "승인" if req.status == "approved" else "피드백 반영 요청"
    try:
        send_push_to_users(
            db, [entry.user_id], title=f"📋 평가 실적 {status_text}",
            body=req.feedback or f"실적이 {status_text}됐어요.", url="/evaluation",
        )
    except Exception as e:
        print(f"[review_actual] 알림 전송 실패: {e}")

    return _serialize_entry(entry)


# ── 등급 부여 ──────────────────────────────────────────────
def _compute_results(db: Session, cycle: EvaluationCycle, current_user: dict) -> dict:
    is_privileged = is_superadmin_email(db, current_user.get("email"))
    if not is_privileged:
        member = _get_member(db, current_user["uid"], cycle.company_id)
        is_privileged = bool(member and member.is_admin)

    entries_q = db.query(EvaluationEntry).filter(EvaluationEntry.cycle_id == cycle.id)
    if not is_privileged:
        entries_q = entries_q.filter(EvaluationEntry.evaluator_id == current_user["uid"])
    entries = entries_q.all()

    by_user: dict[str, list[EvaluationEntry]] = {}
    for e in entries:
        by_user.setdefault(e.user_id, []).append(e)

    results = (
        db.query(EvaluationResult)
        .filter(EvaluationResult.cycle_id == cycle.id, EvaluationResult.user_id.in_(by_user.keys()))
        .all()
        if by_user else []
    )
    result_by_user = {r.user_id: r for r in results}

    dist = cycle.grade_distribution or []
    pool_size = len(by_user)
    target_counts = {g["grade"]: round(pool_size * g.get("ratio", 0) / 100) for g in dist}
    current_counts: dict[str, int] = {}
    for r in result_by_user.values():
        if r.grade:
            current_counts[r.grade] = current_counts.get(r.grade, 0) + 1

    names = _name_map(db, cycle.company_id)
    people = [
        {
            "user_id": uid,
            "user_name": names.get(uid, uid),
            "ready": len(es) >= 2 and all(e.actual_status == "approved" for e in es),
            "score": result_by_user[uid].score if uid in result_by_user else None,
            "grade": result_by_user[uid].grade if uid in result_by_user else None,
        }
        for uid, es in by_user.items()
    ]

    return {
        "people": people,
        "distribution": [
            {
                "grade": g["grade"], "ratio": g.get("ratio", 0),
                "target": target_counts.get(g["grade"], 0), "current": current_counts.get(g["grade"], 0),
            }
            for g in dist
        ],
    }


@router.get("/results/{cycle_id}")
def get_results(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    return _compute_results(db, cycle, current_user)


class GradeRequest(BaseModel):
    cycle_id: str
    score: Optional[float] = None
    grade: str


@router.put("/results/{user_id}/grade")
def set_grade(
    user_id: str, req: GradeRequest,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == req.cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")

    entries = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == req.cycle_id, EvaluationEntry.user_id == user_id,
    ).all()
    if len(entries) < 2 or not all(e.actual_status == "approved" for e in entries):
        raise HTTPException(status_code=400, detail="성과/역량 실적이 모두 승인된 후에 등급을 줄 수 있어요")

    evaluator_id = entries[0].evaluator_id
    if not is_superadmin_email(db, current_user.get("email")) and evaluator_id != current_user["uid"]:
        member = _get_member(db, current_user["uid"], cycle.company_id)
        if not member or not member.is_admin:
            raise HTTPException(status_code=403, detail="담당 평가자 또는 관리자만 등급을 줄 수 있어요")

    valid_grades = {g["grade"] for g in (cycle.grade_distribution or [])}
    if valid_grades and req.grade not in valid_grades:
        raise HTTPException(status_code=400, detail="유효하지 않은 등급이에요")

    result = db.query(EvaluationResult).filter(
        EvaluationResult.cycle_id == req.cycle_id, EvaluationResult.user_id == user_id,
    ).first()
    if not result:
        result = EvaluationResult(
            cycle_id=req.cycle_id, company_id=cycle.company_id, user_id=user_id, evaluator_id=evaluator_id,
        )
        db.add(result)

    result.score = req.score
    result.grade = req.grade
    result.graded_at = datetime.utcnow()
    result.graded_by = current_user["uid"]
    db.commit()

    try:
        send_push_to_users(
            db, [user_id], title="⭐ 평가 등급 확정",
            body=f"{cycle.name} 등급이 확정됐어요.", url="/evaluation",
        )
    except Exception as e:
        print(f"[set_grade] 알림 전송 실패: {e}")

    return {"success": True, "grade": result.grade, "score": result.score}


# ══════════════════════════════════════════════════════════
# Phase 2: AI 분석 (커리어/성장/1on1)
# ══════════════════════════════════════════════════════════

def _serialize_result(r: Optional[EvaluationResult]) -> dict:
    if not r:
        return {
            "score": None, "grade": None,
            "ai_career_analysis": None, "ai_career_generated_at": None,
            "ai_growth_analysis": None, "ai_competencies": [], "ai_growth_generated_at": None,
        }
    return {
        "score": r.score, "grade": r.grade,
        "ai_career_analysis": r.ai_career_analysis,
        "ai_career_generated_at": r.ai_career_generated_at.isoformat() if r.ai_career_generated_at else None,
        "ai_growth_analysis": r.ai_growth_analysis,
        "ai_competencies": r.ai_competencies or [],
        "ai_growth_generated_at": r.ai_growth_generated_at.isoformat() if r.ai_growth_generated_at else None,
    }


def _combined_content(entries: list[EvaluationEntry], phase: str) -> str:
    label = {"performance": "성과평가", "competency": "역량평가"}
    parts = []
    for e in sorted(entries, key=lambda x: x.category):
        content = getattr(e, f"{phase}_content")
        if content and content.strip():
            parts.append(f"[{label.get(e.category, e.category)}]\n{content.strip()}")
    return "\n\n".join(parts)


def _get_or_create_result(db: Session, cycle: EvaluationCycle, user_id: str, evaluator_id: str) -> EvaluationResult:
    result = db.query(EvaluationResult).filter(
        EvaluationResult.cycle_id == cycle.id, EvaluationResult.user_id == user_id,
    ).first()
    if not result:
        result = EvaluationResult(cycle_id=cycle.id, company_id=cycle.company_id, user_id=user_id, evaluator_id=evaluator_id)
        db.add(result)
    return result


# ── 직무 입력 (본인) ───────────────────────────────────────
class JobTitleRequest(BaseModel):
    company_id: str
    job_title: str


@router.put("/job-title")
def set_job_title(
    req: JobTitleRequest, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    member = _require_member(db, current_user["uid"], req.company_id)
    member.job_title = req.job_title.strip() or None
    member.updated_by = current_user["uid"]
    db.commit()
    return {"success": True, "job_title": member.job_title}


# ── 본인 결과(등급/AI분석) 조회 ────────────────────────────
@router.get("/results/me/{cycle_id}")
def get_my_result(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    result = db.query(EvaluationResult).filter(
        EvaluationResult.cycle_id == cycle_id, EvaluationResult.user_id == current_user["uid"],
    ).first()
    return _serialize_result(result)


# ── AI 커리어 분석 (본인, 계획 기반) ────────────────────────
@router.post("/results/{cycle_id}/career-analysis")
def generate_career_analysis(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")

    member = _get_member(db, current_user["uid"], cycle.company_id)
    if not member or not member.job_title:
        raise HTTPException(status_code=400, detail="먼저 직무를 입력해주세요")

    entries = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == cycle_id, EvaluationEntry.user_id == current_user["uid"],
    ).all()
    if not entries:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    plan_text = _combined_content(entries, "plan")
    if not plan_text:
        raise HTTPException(status_code=400, detail="계획을 먼저 작성해주세요")

    try:
        analysis = analyze_career(member.job_title, plan_text)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 분석에 실패했어요: {e}")

    result = _get_or_create_result(db, cycle, current_user["uid"], entries[0].evaluator_id)
    result.ai_career_analysis = analysis
    result.ai_career_generated_at = datetime.utcnow()
    db.commit()
    return _serialize_result(result)


# ── AI 성장 분석 + 역량 레이더 (본인, 실적 기반) ────────────
@router.post("/results/{cycle_id}/growth-analysis")
def generate_growth_analysis(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")

    member = _get_member(db, current_user["uid"], cycle.company_id)
    if not member or not member.job_title:
        raise HTTPException(status_code=400, detail="먼저 직무를 입력해주세요")

    entries = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == cycle_id, EvaluationEntry.user_id == current_user["uid"],
    ).all()
    if not entries:
        raise HTTPException(status_code=404, detail="평가 항목을 찾을 수 없어요")
    plan_text = _combined_content(entries, "plan")
    actual_text = _combined_content(entries, "actual")
    if not actual_text:
        raise HTTPException(status_code=400, detail="실적을 먼저 작성해주세요")

    try:
        analysis, competencies = analyze_growth(member.job_title, plan_text, actual_text)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 분석에 실패했어요: {e}")

    result = _get_or_create_result(db, cycle, current_user["uid"], entries[0].evaluator_id)
    result.ai_growth_analysis = analysis
    result.ai_competencies = competencies
    result.ai_growth_generated_at = datetime.utcnow()
    db.commit()
    return _serialize_result(result)


# ── 1on1 면담 녹음/분석 ─────────────────────────────────────
def _is_one_on_one_viewer(db: Session, current_user: dict, company_id: str, evaluator_id: str) -> bool:
    """평가관리자(회사 관리자 또는 평가자 소속팀의 상위팀 관리자)만 1on1 분석을 열람할 수 있다."""
    if is_superadmin_email(db, current_user.get("email")):
        return True
    member = _get_member(db, current_user["uid"], company_id)
    if member and member.is_admin:
        return True

    evaluator_team_ids = [
        row.team_id for row in db.query(TeamMember.team_id).filter(TeamMember.user_id == evaluator_id).all()
    ]
    if not evaluator_team_ids:
        return False
    parent_team_ids = {
        t.parent_team_id for t in db.query(Team).filter(Team.id.in_(evaluator_team_ids)).all() if t.parent_team_id
    }
    if not parent_team_ids:
        return False
    upper_manager = db.query(Team).filter(
        Team.id.in_(parent_team_ids), Team.manager_id == current_user["uid"],
    ).first()
    return upper_manager is not None


def _process_one_on_one_background(session_id: str, audio_bytes: bytes, filename: str, content_type: str | None):
    db = SessionLocal()
    try:
        session = db.query(OneOnOneSession).filter(OneOnOneSession.id == session_id).first()
        if not session:
            return
        try:
            transcript = transcribe_audio(audio_bytes, filename, content_type)
            if not transcript:
                raise ValueError("음성에서 텍스트를 인식하지 못했어요")
            analysis = analyze_one_on_one(transcript)

            session.transcript = transcript
            session.ai_analysis = analysis
            session.status = "completed"
            db.commit()
        except Exception as e:
            print(f"[_process_one_on_one_background] 처리 실패: {e}")
            db.rollback()
            session.status = "failed"
            session.error_message = str(e)[:500]
            db.commit()
    finally:
        db.close()


@router.post("/one-on-one/upload")
async def upload_one_on_one(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    cycle_id: str = Form(...),
    evaluatee_user_id: str = Form(...),
    duration_seconds: float = Form(0),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")

    entry = db.query(EvaluationEntry).filter(
        EvaluationEntry.cycle_id == cycle_id, EvaluationEntry.user_id == evaluatee_user_id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="평가 대상을 찾을 수 없어요")
    if entry.evaluator_id != current_user["uid"]:
        member = _get_member(db, current_user["uid"], cycle.company_id)
        if not member or not member.is_admin:
            raise HTTPException(status_code=403, detail="담당 평가자 또는 관리자만 녹음할 수 있어요")

    audio_bytes = await file.read()
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="녹음 파일이 비어있어요")

    session = OneOnOneSession(
        cycle_id=cycle_id, company_id=cycle.company_id,
        evaluator_id=entry.evaluator_id, evaluatee_id=evaluatee_user_id,
        duration_seconds=duration_seconds or None, status="processing",
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    background_tasks.add_task(
        _process_one_on_one_background, session.id, audio_bytes, file.filename or "recording.webm", file.content_type,
    )
    return {"success": True, "id": session.id, "status": "processing"}


def _serialize_one_on_one_sessions(db: Session, cycle: EvaluationCycle, current_user: dict) -> list[dict]:
    sessions = db.query(OneOnOneSession).filter(OneOnOneSession.cycle_id == cycle.id).all()
    visible = [s for s in sessions if _is_one_on_one_viewer(db, current_user, cycle.company_id, s.evaluator_id)]
    if not visible and sessions:
        # 세션은 있지만 열람 권한이 없는 경우와, 세션 자체가 없는 경우를 구분
        raise HTTPException(status_code=403, detail="평가관리자만 열람할 수 있어요")

    names = _name_map(db, cycle.company_id)
    return [
        {
            "id": s.id,
            "evaluator_name": names.get(s.evaluator_id, s.evaluator_id),
            "evaluatee_name": names.get(s.evaluatee_id, s.evaluatee_id),
            "status": s.status,
            "ai_analysis": s.ai_analysis,
            "recorded_at": s.recorded_at.isoformat(),
            "duration_seconds": s.duration_seconds,
        }
        for s in visible
    ]


@router.get("/one-on-one/{cycle_id}")
def list_one_on_one(
    cycle_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    cycle = db.query(EvaluationCycle).filter(EvaluationCycle.id == cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="평가 코드를 찾을 수 없어요")
    return {"sessions": _serialize_one_on_one_sessions(db, cycle, current_user)}


@router.get("/bootstrap/one-on-one")
def get_bootstrap_one_on_one(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """/evaluation/one-on-one/monitor 화면용. company/my + cycles/active + one-on-one 목록을
    하나로 묶어 요청 왕복 수를 3회 → 1회로 줄인다."""
    uid = current_user["uid"]
    member = _get_primary_member(db, uid)
    company_id = member.company_id if member else None
    if not company_id:
        return {"company_id": None, "sessions": []}

    cycle_row = _get_active_cycle_row(db, company_id)
    if not cycle_row:
        return {"company_id": company_id, "sessions": []}

    return {"company_id": company_id, "sessions": _serialize_one_on_one_sessions(db, cycle_row, current_user)}
