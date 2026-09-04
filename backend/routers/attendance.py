from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from typing import Optional
import os
from sqlalchemy.orm import Session
from sqlalchemy import func
from database.connection import get_db
from models.attendance import Attendance
from models.location import Location
from models.company import Company, CompanyMember
from models.reclock import ReclockRequest
from routers.deps import get_current_user
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
from utils.workday import get_work_day_range, kst_date_str as _kst_date_str, today_kst_str
from utils.admin import is_superadmin_email

router = APIRouter()

KST = timezone(timedelta(hours=9))


def _approved_reclock_minutes_map(db: Session, user_id: str, date_strs: list[str]) -> dict:
    """승인된 재출근 세션의 근무 분(分)을 work_date별로 합산."""
    if not date_strs:
        return {}
    rows = (
        db.query(ReclockRequest)
        .filter(
            ReclockRequest.user_id == user_id,
            ReclockRequest.status == "approved",
            ReclockRequest.work_date.in_(date_strs),
            ReclockRequest.checkout_at.isnot(None),
        )
        .all()
    )
    result: dict = {}
    for r in rows:
        mins = max(0, int((r.checkout_at - r.checkin_at).total_seconds() / 60))
        result[r.work_date] = result.get(r.work_date, 0) + mins
    return result


def _approved_reclock_minutes(db: Session, user_id: str, date_str: str) -> int:
    return _approved_reclock_minutes_map(db, user_id, [date_str]).get(date_str, 0)

@router.get("/summary/{user_id}")
def get_attendance_summary(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["uid"] != user_id:
        raise HTTPException(status_code=403, detail="본인의 기록만 조회할 수 있어요")

    start, end = get_work_day_range()
    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= start,
            Attendance.recorded_at < end,
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    checkin = next((r for r in records if r.type == "checkin"), None)
    checkout = next((r for r in records if r.type == "checkout"), None)

    outing_minutes = (checkin.outing_minutes or 0) if checkin else 0

    work_minutes = 0
    if checkin and checkout:
        diff = checkout.recorded_at - checkin.recorded_at
        work_minutes = max(0, int(diff.total_seconds() / 60) - outing_minutes)

    work_minutes += _approved_reclock_minutes(db, user_id, today_kst_str())

    return {
        "date": start.date().isoformat(),
        "checkin": checkin.recorded_at.isoformat() if checkin else None,
        "checkout": checkout.recorded_at.isoformat() if checkout else None,
        "checkin_address": checkin.address if checkin else None,
        "checkout_address": checkout.address if checkout else None,
        "is_remote": bool(checkin.is_remote) if checkin else False,
        "is_outing": bool(checkin.is_outing) if checkin else False,
        "outing_minutes": outing_minutes,
        "work_minutes": work_minutes,
        "work_hours": f"{work_minutes // 60}시간 {work_minutes % 60}분",
    }


@router.get("/company/{company_id}")
def get_company_attendance(
    company_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    is_superadmin = is_superadmin_email(db, current_user.get("email"))
    requester = db.query(CompanyMember).filter(
        CompanyMember.user_id == current_user["uid"],
        CompanyMember.company_id == company_id,
        CompanyMember.is_admin == True,
    ).first()
    if not requester and not is_superadmin:
        raise HTTPException(status_code=403, detail="관리자만 근태 현황을 조회할 수 있어요")

    start, end = get_work_day_range()
    now = datetime.now(KST)

    members = db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    if not members:
        return {"attendance": []}

    user_ids = [m.user_id for m in members]
    all_records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id.in_(user_ids),
            Attendance.recorded_at >= start,
            Attendance.recorded_at < end,
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    records_by_user: dict = {}
    for r in all_records:
        records_by_user.setdefault(r.user_id, []).append(r)

    today_str = today_kst_str()
    reclock_today = (
        db.query(ReclockRequest)
        .filter(
            ReclockRequest.user_id.in_(user_ids),
            ReclockRequest.work_date == today_str,
            ReclockRequest.status.in_(["in_progress", "pending", "approved"]),
        )
        .order_by(ReclockRequest.created_at.desc())
        .all()
    )
    reclock_by_user: dict = {}
    for r in reclock_today:
        reclock_by_user.setdefault(r.user_id, []).append(r)

    result = []
    for member in members:
        records = records_by_user.get(member.user_id, [])
        checkin = next((r for r in records if r.type == "checkin"), None)
        checkout = next((r for r in records if r.type == "checkout"), None)

        is_missing_checkout = (
            checkin is not None
            and checkout is None
            and now >= end
        )

        work_minutes = 0
        if checkin and checkout:
            diff = checkout.recorded_at - checkin.recorded_at
            work_minutes = max(0, int(diff.total_seconds() / 60) - (checkin.outing_minutes or 0))

        member_reclock = reclock_by_user.get(member.user_id, [])
        for r in member_reclock:
            if r.status == "approved" and r.checkout_at:
                work_minutes += max(0, int((r.checkout_at - r.checkin_at).total_seconds() / 60))
        reclock_status = next((r.status for r in member_reclock if r.status in ("in_progress", "pending")), None)

        if checkin and not checkout and not is_missing_checkout:
            status = "출근중"
        elif checkout:
            status = "퇴근"
        elif is_missing_checkout:
            status = "미퇴근"
        else:
            status = "미출근"

        result.append({
            "user_id": member.user_id,
            "user_name": member.user_name,
            "user_email": member.user_email,
            "checkin": checkin.recorded_at.isoformat() if checkin else None,
            "checkout": checkout.recorded_at.isoformat() if checkout else None,
            "work_hours": f"{work_minutes // 60}h {work_minutes % 60}m" if work_minutes else "-",
            "status": status,
            "is_missing_checkout": is_missing_checkout,
            "is_outing": bool(checkin.is_outing) if checkin and not checkout else False,
            "reclock_status": reclock_status,
        })

    return {"attendance": result}


def compute_weekly_report(db: Session, user_id: str, week_start=None) -> dict:
    """주간 근무 리포트(승인된 재출근 포함)를 계산. week_start가 없으면 이번 주.
    /weekly 엔드포인트와 52시간 경고 스케줄러가 함께 사용."""
    today = datetime.now(KST).date()
    if week_start is None:
        week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    # KST 기준 주간 범위를 UTC로 변환하여 필터
    utc_start = datetime(week_start.year, week_start.month, week_start.day, 0, 0, 0) - timedelta(hours=9)
    utc_end = datetime(week_end.year, week_end.month, week_end.day, 23, 59, 59) - timedelta(hours=9)

    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= utc_start,
            Attendance.recorded_at <= utc_end,
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    daily = {}
    for r in records:
        # KST 기준으로 날짜 버킷팅 (UTC 저장이므로 +9h 변환)
        kst_time = r.recorded_at.replace(tzinfo=timezone.utc).astimezone(KST)
        date_str = kst_time.date().isoformat()
        if date_str not in daily:
            daily[date_str] = {"checkin": None, "checkout": None, "work_minutes": 0, "outing_minutes": 0}
        if r.type == "checkin" and not daily[date_str]["checkin"]:
            daily[date_str]["checkin"] = r.recorded_at.isoformat()
            daily[date_str]["checkin_address"] = r.address
            daily[date_str]["outing_minutes"] = r.outing_minutes or 0
        if r.type == "checkout":
            daily[date_str]["checkout"] = r.recorded_at.isoformat()
            daily[date_str]["checkout_address"] = r.address

    now_kst = datetime.now(KST)
    today_str = now_kst.date().isoformat()

    for date_str, data in daily.items():
        if data["checkin"] and data["checkout"]:
            checkin = datetime.fromisoformat(data["checkin"])
            checkout = datetime.fromisoformat(data["checkout"])
            diff = checkout - checkin
            data["work_minutes"] = max(0, int(diff.total_seconds() / 60) - data.get("outing_minutes", 0))
            data["work_hours"] = f"{data['work_minutes'] // 60}시간 {data['work_minutes'] % 60}분"
        elif data["checkin"] and not data["checkout"] and date_str == today_str:
            # 오늘 퇴근 전: 현재 시각 기준으로 합산 (완료된 외출 시간만 차감, 진행 중인 외출은 반영 안 함)
            checkin = datetime.fromisoformat(data["checkin"])
            if checkin.tzinfo is None:
                checkin = checkin.replace(tzinfo=timezone.utc)
            diff = now_kst - checkin.astimezone(KST)
            data["work_minutes"] = max(0, int(diff.total_seconds() / 60) - data.get("outing_minutes", 0))
            data["work_hours"] = f"{data['work_minutes'] // 60}시간 {data['work_minutes'] % 60}분 (진행중)"
        else:
            data["work_hours"] = "-"

    reclock_map = _approved_reclock_minutes_map(db, user_id, list(daily.keys()))
    for date_str, extra_minutes in reclock_map.items():
        if extra_minutes <= 0:
            continue
        bucket = daily.setdefault(date_str, {"checkin": None, "checkout": None, "work_minutes": 0, "outing_minutes": 0})
        bucket["work_minutes"] += extra_minutes
        bucket["work_hours"] = f"{bucket['work_minutes'] // 60}시간 {bucket['work_minutes'] % 60}분"

    total_minutes = sum(d["work_minutes"] for d in daily.values())

    overtime_52h = total_minutes > 52 * 60

    return {
        "user_id": user_id,
        "period": f"{week_start} ~ {week_end}",
        "week_start": str(week_start),
        "total_work_hours": f"{total_minutes // 60}시간 {total_minutes % 60}분",
        "total_minutes": total_minutes,
        "overtime_52h": overtime_52h,
        "daily": daily,
    }


def compute_month_to_date_minutes(db: Session, user_id: str, end_date) -> int:
    """이번 달 1일부터 end_date까지(포함) 근무 분 합계(승인된 재출근 포함).
    근로시간 패턴 알림 스케줄러가 월 누적치 계산에 사용."""
    month_start = end_date.replace(day=1)
    utc_start = datetime(month_start.year, month_start.month, month_start.day, 0, 0, 0) - timedelta(hours=9)
    utc_end = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59) - timedelta(hours=9)

    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= utc_start,
            Attendance.recorded_at <= utc_end,
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    daily: dict = {}
    for r in records:
        kst_time = r.recorded_at.replace(tzinfo=timezone.utc).astimezone(KST)
        date_str = kst_time.date().isoformat()
        bucket = daily.setdefault(date_str, {"checkin": None, "checkout": None, "outing_minutes": 0})
        if r.type == "checkin" and not bucket["checkin"]:
            bucket["checkin"] = r.recorded_at
            bucket["outing_minutes"] = r.outing_minutes or 0
        if r.type == "checkout":
            bucket["checkout"] = r.recorded_at

    total_minutes = 0
    for data in daily.values():
        if data["checkin"] and data["checkout"]:
            diff = data["checkout"] - data["checkin"]
            total_minutes += max(0, int(diff.total_seconds() / 60) - data["outing_minutes"])

    reclock_map = _approved_reclock_minutes_map(db, user_id, list(daily.keys()))
    total_minutes += sum(reclock_map.values())

    return total_minutes


class WorkHourLimitsRequest(BaseModel):
    max_weekly_hours: int
    max_monthly_hours: Optional[int] = None


@router.put("/work-hour-limits/{company_id}")
def set_work_hour_limits(
    company_id: str,
    req: WorkHourLimitsRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없어요")

    is_superadmin = is_superadmin_email(db, current_user.get("email"))
    member = db.query(CompanyMember).filter(
        CompanyMember.user_id == current_user["uid"],
        CompanyMember.company_id == company_id,
        CompanyMember.is_admin == True,
    ).first()
    if not member and not is_superadmin:
        raise HTTPException(status_code=403, detail="관리자만 설정할 수 있어요")

    if req.max_weekly_hours <= 0:
        raise HTTPException(status_code=400, detail="주 최대 근로시간은 0보다 커야 해요")
    if req.max_monthly_hours is not None and req.max_monthly_hours <= 0:
        raise HTTPException(status_code=400, detail="월 최대 근로시간은 0보다 커야 해요")

    company.max_weekly_minutes = req.max_weekly_hours * 60
    company.max_monthly_minutes = req.max_monthly_hours * 60 if req.max_monthly_hours else None
    company.updated_by = current_user["uid"]
    db.commit()
    return {
        "success": True,
        "max_weekly_hours": req.max_weekly_hours,
        "max_monthly_hours": req.max_monthly_hours,
    }


@router.get("/weekly/{user_id}")
def get_weekly_report(
    user_id: str,
    start_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["uid"] != user_id:
        raise HTTPException(status_code=403, detail="본인의 기록만 조회할 수 있어요")

    week_start = None
    if start_date:
        try:
            week_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            week_start = None

    return compute_weekly_report(db, user_id, week_start)


@router.get("/monthly/{user_id}")
def get_monthly_report(
    user_id: str,
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    import calendar as _cal

    if current_user["uid"] != user_id:
        raise HTTPException(status_code=403, detail="본인의 기록만 조회할 수 있어요")

    today = datetime.now(KST).date()
    target_year = year if year else today.year
    target_month = month if month else today.month

    _, last_day = _cal.monthrange(target_year, target_month)
    month_start = datetime(target_year, target_month, 1).date()
    month_end = datetime(target_year, target_month, last_day).date()

    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= month_start,
            Attendance.recorded_at < month_end + timedelta(days=1),
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    daily = {}
    for r in records:
        date_str = _kst_date_str(r.recorded_at)
        if date_str not in daily:
            daily[date_str] = {"checkin": None, "checkout": None, "work_minutes": 0}
        if r.type == "checkin" and not daily[date_str]["checkin"]:
            daily[date_str]["checkin"] = r.recorded_at.isoformat()
        if r.type == "checkout":
            daily[date_str]["checkout"] = r.recorded_at.isoformat()

    for date_str, data in daily.items():
        if data["checkin"] and data["checkout"]:
            checkin = datetime.fromisoformat(data["checkin"])
            checkout = datetime.fromisoformat(data["checkout"])
            diff = checkout - checkin
            data["work_minutes"] = max(0, int(diff.total_seconds() / 60))
            data["work_hours"] = f"{data['work_minutes'] // 60}시간 {data['work_minutes'] % 60}분"
        else:
            data["work_hours"] = "-"

    reclock_map = _approved_reclock_minutes_map(db, user_id, list(daily.keys()))
    for date_str, extra_minutes in reclock_map.items():
        if extra_minutes <= 0:
            continue
        bucket = daily.setdefault(date_str, {"checkin": None, "checkout": None, "work_minutes": 0})
        bucket["work_minutes"] += extra_minutes
        bucket["work_hours"] = f"{bucket['work_minutes'] // 60}시간 {bucket['work_minutes'] % 60}분"

    total_minutes = sum(d["work_minutes"] for d in daily.values())
    work_days = len([d for d in daily.values() if d["work_minutes"] > 0])

    return {
        "user_id": user_id,
        "period": f"{target_year}년 {target_month}월",
        "year": target_year,
        "month": target_month,
        "total_work_hours": f"{total_minutes // 60}시간 {total_minutes % 60}분",
        "work_days": work_days,
        "avg_work_hours": (
            f"{(total_minutes // work_days) // 60}시간 {(total_minutes // work_days) % 60}분"
            if work_days > 0 else "-"
        ),
        "daily": daily,
    }


@router.get("/company-report/{company_id}")
def get_company_report(
    company_id: str,
    type: str = Query(default="weekly"),
    start_date: Optional[str] = Query(default=None),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    import calendar as _cal

    is_superadmin = is_superadmin_email(db, current_user.get("email"))
    requester = db.query(CompanyMember).filter(
        CompanyMember.user_id == current_user["uid"],
        CompanyMember.company_id == company_id,
    ).first()
    if not requester and not is_superadmin:
        raise HTTPException(status_code=403, detail="해당 기업의 멤버만 조회할 수 있어요")

    today = datetime.now(KST).date()
    period_label = ""

    if type == "weekly":
        if start_date:
            try:
                period_start = datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                period_start = today - timedelta(days=today.weekday())
        else:
            period_start = today - timedelta(days=today.weekday())
        period_end = period_start + timedelta(days=6)
        period_label = f"{period_start} ~ {period_end}"
    else:
        target_year = year if year else today.year
        target_month = month if month else today.month
        _, last_day = _cal.monthrange(target_year, target_month)
        period_start = datetime(target_year, target_month, 1).date()
        period_end = datetime(target_year, target_month, last_day).date()
        period_label = f"{target_year}년 {target_month}월"

    members = db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    user_ids = [m.user_id for m in members]

    all_records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id.in_(user_ids),
            Attendance.recorded_at >= period_start,
            Attendance.recorded_at < period_end + timedelta(days=1),
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    records_by_user: dict = {}
    for r in all_records:
        records_by_user.setdefault(r.user_id, []).append(r)

    all_reclock = (
        db.query(ReclockRequest)
        .filter(
            ReclockRequest.user_id.in_(user_ids),
            ReclockRequest.status == "approved",
            ReclockRequest.work_date >= str(period_start),
            ReclockRequest.work_date <= str(period_end),
            ReclockRequest.checkout_at.isnot(None),
        )
        .all()
    )
    reclock_by_user: dict = {}
    for r in all_reclock:
        mins = max(0, int((r.checkout_at - r.checkin_at).total_seconds() / 60))
        per_date = reclock_by_user.setdefault(r.user_id, {})
        per_date[r.work_date] = per_date.get(r.work_date, 0) + mins

    result = []
    for member in members:
        recs = records_by_user.get(member.user_id, [])
        daily: dict = {}
        for r in recs:
            date_str = _kst_date_str(r.recorded_at)
            if date_str not in daily:
                daily[date_str] = {"checkin": None, "checkout": None, "work_minutes": 0}
            if r.type == "checkin" and not daily[date_str]["checkin"]:
                daily[date_str]["checkin"] = r.recorded_at.isoformat()
            if r.type == "checkout":
                daily[date_str]["checkout"] = r.recorded_at.isoformat()

        total_minutes = 0
        work_days = 0
        for d in daily.values():
            if d["checkin"] and d["checkout"]:
                diff = datetime.fromisoformat(d["checkout"]) - datetime.fromisoformat(d["checkin"])
                mins = max(0, int(diff.total_seconds() / 60))
                d["work_minutes"] = mins
                total_minutes += mins
                work_days += 1

        for date_str, extra_minutes in reclock_by_user.get(member.user_id, {}).items():
            bucket = daily.setdefault(date_str, {"checkin": None, "checkout": None, "work_minutes": 0})
            bucket["work_minutes"] += extra_minutes
            total_minutes += extra_minutes

        result.append({
            "user_id": member.user_id,
            "user_name": member.user_name,
            "user_email": member.user_email,
            "work_days": work_days,
            "total_work_hours": f"{total_minutes // 60}시간 {total_minutes % 60}분",
            "total_minutes": total_minutes,
            "overtime_52h": type == "weekly" and total_minutes > 52 * 60,
            "daily": daily,
        })

    return {
        "period": period_label,
        "period_start": str(period_start),
        "period_end": str(period_end),
        "members": result,
    }


@router.delete("/reset/{user_id}")
def reset_attendance(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from models.attendance_reset_log import AttendanceResetLog

    is_superadmin = is_superadmin_email(db, current_user.get("email"))

    target_member = db.query(CompanyMember).filter(CompanyMember.user_id == user_id).first()

    is_self = current_user["uid"] == user_id
    if not is_self and not is_superadmin:
        requester = db.query(CompanyMember).filter(
            CompanyMember.user_id == current_user["uid"],
            CompanyMember.company_id == target_member.company_id if target_member else None,
        ).first()
        if not requester or not requester.is_admin:
            raise HTTPException(status_code=403, detail="본인 또는 소속 회사 관리자만 초기화할 수 있어요")

    role = "self" if is_self else ("superadmin" if is_superadmin else "admin")

    performer_member = db.query(CompanyMember).filter(CompanyMember.user_id == current_user["uid"]).first()
    performer_name = (performer_member.user_name if performer_member else None) or current_user.get("name") or current_user.get("email")

    start, end = get_work_day_range()

    attendance_count = db.query(Attendance).filter(
        Attendance.user_id == user_id,
        Attendance.recorded_at >= start,
        Attendance.recorded_at < end,
    ).delete()

    location_count = db.query(Location).filter(
        Location.user_id == user_id,
        Location.recorded_at >= start,
        Location.recorded_at < end,
    ).delete()

    reclock_count = db.query(ReclockRequest).filter(
        ReclockRequest.user_id == user_id,
        ReclockRequest.work_date == today_kst_str(),
    ).delete()

    db.add(AttendanceResetLog(
        company_id=target_member.company_id if target_member else None,
        target_user_id=user_id,
        target_user_name=target_member.user_name if target_member else None,
        performed_by=current_user["uid"],
        performed_by_name=performer_name,
        performed_by_role=role,
        reset_date=today_kst_str(),
        attendance_count=attendance_count,
        location_count=location_count,
        reclock_count=reclock_count,
    ))

    db.commit()
    return {"message": "오늘 기록 초기화 완료"}


@router.get("/reset-logs/{company_id}")
def get_reset_logs(
    company_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    from models.attendance_reset_log import AttendanceResetLog

    is_superadmin = is_superadmin_email(db, current_user.get("email"))
    requester = db.query(CompanyMember).filter(
        CompanyMember.user_id == current_user["uid"],
        CompanyMember.company_id == company_id,
        CompanyMember.is_admin == True,
    ).first()
    if not requester and not is_superadmin:
        raise HTTPException(status_code=403, detail="관리자만 조회할 수 있어요")

    rows = (
        db.query(AttendanceResetLog)
        .filter(AttendanceResetLog.company_id == company_id)
        .order_by(AttendanceResetLog.created_at.desc())
        .limit(200)
        .all()
    )
    return {
        "logs": [
            {
                "id": r.id,
                "target_user_id": r.target_user_id,
                "target_user_name": r.target_user_name,
                "performed_by": r.performed_by,
                "performed_by_name": r.performed_by_name,
                "performed_by_role": r.performed_by_role,
                "reset_date": r.reset_date,
                "attendance_count": r.attendance_count,
                "location_count": r.location_count,
                "reclock_count": r.reclock_count,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]
    }


from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io


@router.get("/export/{company_id}")
def export_attendance_excel(
    company_id: str,
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    from models.company import Company, CompanyMember
    import calendar

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없습니다")

    is_superadmin = is_superadmin_email(db, current_user.get("email"))
    requester = db.query(CompanyMember).filter(
        CompanyMember.user_id == current_user["uid"],
        CompanyMember.company_id == company_id,
        CompanyMember.is_admin == True,
    ).first()
    if not requester and not is_superadmin:
        raise HTTPException(status_code=403, detail="관리자만 내보내기 할 수 있어요")

    members = (
        db.query(CompanyMember).filter(CompanyMember.company_id == company_id).all()
    )

    now_kst = datetime.now(KST)
    target_year = year if year else now_kst.year
    target_month = month  # None이면 연도 전체

    if target_month:
        start_date = datetime(target_year, target_month, 1).date()
        last_day = calendar.monthrange(target_year, target_month)[1]
        end_date = datetime(target_year, target_month, last_day).date()
    else:
        start_date = datetime(target_year, 1, 1).date()
        end_date = datetime(target_year, 12, 31).date()

    user_ids = [m.user_id for m in members]
    all_records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id.in_(user_ids),
            Attendance.recorded_at >= datetime.combine(start_date, datetime.min.time()),
            Attendance.recorded_at <= datetime.combine(end_date, datetime.max.time()),
        )
        .order_by(Attendance.recorded_at)
        .all()
    )
    records_by_user: dict = {}
    for r in all_records:
        records_by_user.setdefault(r.user_id, []).append(r)

    all_reclock = (
        db.query(ReclockRequest)
        .filter(
            ReclockRequest.user_id.in_(user_ids),
            ReclockRequest.status == "approved",
            ReclockRequest.work_date >= start_date.isoformat(),
            ReclockRequest.work_date <= end_date.isoformat(),
            ReclockRequest.checkout_at.isnot(None),
        )
        .all()
    )
    # 팀장 최종 승인(approved)된 재출근 건만 공식 기록으로 반영 — 하루 최대 1건
    reclock_by_user: dict = {}
    for r in all_reclock:
        per_date = reclock_by_user.setdefault(r.user_id, {})
        per_date[r.work_date] = r

    def _reclock_minutes(r):
        if not r or not r.checkout_at:
            return 0
        return max(0, int((r.checkout_at - r.checkin_at).total_seconds() / 60))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "근무기록"

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "이름", "이메일", "날짜", "출근시간", "퇴근시간", "근무시간(분)", "출근위치", "퇴근위치",
        "재출근시작시간", "재출근시작주소", "재출근종료시간", "재출근종료주소", "재출근 총 시간(분)",
        "총 외출 시간(분)", "주차", "주간근무시간(분)", "52시간초과",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    def get_week_key(d):
        monday = d - timedelta(days=d.weekday())
        return monday.isoformat()

    row = 2
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    total_cols = len(headers)

    for member in members:
        records = records_by_user.get(member.user_id, [])
        member_reclock = reclock_by_user.get(member.user_id, {})

        daily = {}
        for r in records:
            date_str = _kst_date_str(r.recorded_at)
            if date_str not in daily:
                daily[date_str] = {"checkin": None, "checkout": None}
            if r.type == "checkin" and not daily[date_str]["checkin"]:
                daily[date_str]["checkin"] = r
            if r.type == "checkout":
                daily[date_str]["checkout"] = r
        for date_str in member_reclock:
            daily.setdefault(date_str, {"checkin": None, "checkout": None})

        # 주차별 합산 (승인된 재출근 시간 포함, 외출 시간 제외)
        weekly_minutes: dict = {}
        for date_str, data in daily.items():
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            wk = get_week_key(d)
            checkin = data["checkin"]
            checkout = data["checkout"]
            if checkin and checkout:
                mins = max(0, int((checkout.recorded_at - checkin.recorded_at).total_seconds() / 60) - (checkin.outing_minutes or 0))
            else:
                mins = 0
            mins += _reclock_minutes(member_reclock.get(date_str))
            weekly_minutes[wk] = weekly_minutes.get(wk, 0) + mins

        for date_str, data in sorted(daily.items()):
            checkin = data["checkin"]
            checkout = data["checkout"]
            reclock = member_reclock.get(date_str)
            reclock_minutes = _reclock_minutes(reclock)
            outing_minutes = (checkin.outing_minutes or 0) if checkin else 0

            work_minutes = 0
            if checkin and checkout:
                diff = checkout.recorded_at - checkin.recorded_at
                work_minutes = max(0, int(diff.total_seconds() / 60) - outing_minutes)
            work_minutes += reclock_minutes

            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            wk = get_week_key(d)
            week_total = weekly_minutes.get(wk, 0)
            is_overtime = week_total > 52 * 60

            ws.cell(row=row, column=1, value=member.user_name or "-")
            ws.cell(row=row, column=2, value=member.user_email)
            ws.cell(row=row, column=3, value=date_str)
            ws.cell(row=row, column=4, value=checkin.recorded_at.strftime("%H:%M") if checkin else "-")
            ws.cell(row=row, column=5, value=checkout.recorded_at.strftime("%H:%M") if checkout else "-")
            ws.cell(row=row, column=6, value=work_minutes if work_minutes > 0 else "-")
            ws.cell(row=row, column=7, value=checkin.address if checkin else "-")
            ws.cell(row=row, column=8, value=checkout.address if checkout else "-")
            ws.cell(row=row, column=9, value=reclock.checkin_at.strftime("%H:%M") if reclock and reclock.checkin_at else "-")
            ws.cell(row=row, column=10, value=reclock.checkin_address if reclock and reclock.checkin_address else "-")
            ws.cell(row=row, column=11, value=reclock.checkout_at.strftime("%H:%M") if reclock and reclock.checkout_at else "-")
            ws.cell(row=row, column=12, value=reclock.checkout_address if reclock and reclock.checkout_address else "-")
            ws.cell(row=row, column=13, value=reclock_minutes if reclock_minutes > 0 else "-")
            ws.cell(row=row, column=14, value=outing_minutes if outing_minutes > 0 else "-")
            ws.cell(row=row, column=15, value=f"{wk} 주")
            ws.cell(row=row, column=16, value=week_total if week_total > 0 else "-")
            ws.cell(row=row, column=17, value="초과" if is_overtime else "-")
            if is_overtime:
                for col in range(1, total_cols + 1):
                    ws.cell(row=row, column=col).fill = red_fill
            row += 1

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    if target_month:
        period_label = f"{target_year}년{target_month}월"
    else:
        period_label = f"{target_year}년_전체"
    filename = f"{company.name}_근무기록_{period_label}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/month/{user_id}")
def get_month_dates(
    user_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["uid"] != user_id:
        raise HTTPException(status_code=403, detail="본인의 기록만 조회할 수 있어요")

    start = datetime(year, month, 1, tzinfo=KST)
    if month == 12:
        end = datetime(year + 1, 1, 1, tzinfo=KST)
    else:
        end = datetime(year, month + 1, 1, tzinfo=KST)

    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= start,
            Attendance.recorded_at < end,
        )
        .all()
    )

    dates = {}
    for r in records:
        date_str = r.recorded_at.date().isoformat()
        dates[date_str] = True

    return {"dates": dates}


@router.get("/day/{user_id}")
def get_day_record(
    user_id: str,
    date: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["uid"] != user_id:
        raise HTTPException(status_code=403, detail="본인의 기록만 조회할 수 있어요")

    try:
        target = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=KST)
    except ValueError:
        return {"error": "날짜 형식 오류"}

    start = target + timedelta(hours=4)
    end = start + timedelta(hours=24)

    records = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.recorded_at >= start,
            Attendance.recorded_at < end,
        )
        .order_by(Attendance.recorded_at)
        .all()
    )

    checkin = next((r for r in records if r.type == "checkin"), None)
    checkout = next((r for r in records if r.type == "checkout"), None)

    work_minutes = None
    if checkin and checkout:
        diff = checkout.recorded_at - checkin.recorded_at
        work_minutes = int(diff.total_seconds() / 60)

    return {
        "date": date,
        "checkin": checkin.recorded_at.isoformat() if checkin else None,
        "checkout": checkout.recorded_at.isoformat() if checkout else None,
        "checkin_address": checkin.address if checkin else None,
        "checkout_address": checkout.address if checkout else None,
        "work_minutes": work_minutes,
    }